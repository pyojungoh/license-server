"""
엑셀 파일 읽기 모듈
엑셀 파일에서 송장번호를 읽어옵니다.
"""

from openpyxl import load_workbook
from typing import List
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelReader:
    """엑셀 파일에서 송장번호 읽기 (헤더 없이 1행 1열부터 읽기)"""
    
    def __init__(self, file_path: str, sheet_name: str = "Sheet1"):
        """
        초기화
        
        Args:
            file_path: 엑셀 파일 경로
            sheet_name: 시트 이름
        """
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
    
    def read_invoices(self) -> List[str]:
        """
        엑셀에서 송장번호 읽기 (1행 1열부터 헤더 없이 읽기)
        
        Returns:
            송장번호 리스트 (문자열로 변환하여 앞자리 0 보존)
            
        Raises:
            FileNotFoundError: 파일을 찾을 수 없을 때
            ValueError: 시트를 찾을 수 없을 때
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {self.file_path}")
        
        logger.info(f"엑셀 파일 읽기: {self.file_path}")
        
        try:
            # 엑셀 파일 열기
            wb = load_workbook(self.file_path, data_only=True)
            
            # 시트 선택
            if self.sheet_name not in wb.sheetnames:
                raise ValueError(f"시트 '{self.sheet_name}'을 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}")
            
            ws = wb[self.sheet_name]
            logger.info(f"시트 선택: {self.sheet_name}")
            
            # 1행 1열부터 데이터 읽기 (헤더 없이)
            invoices = []
            for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                # 첫 번째 열(A열)의 값만 읽기
                value = row[0] if len(row) > 0 else None
                
                # 값이 있으면 문자열로 변환 (앞자리 0 보존)
                if value is not None:
                    invoice_str = str(value).strip()
                    if invoice_str:  # 빈 문자열이 아니면
                        invoices.append(invoice_str)
                        logger.debug(f"행 {row_num}: {invoice_str}")
            
            logger.info(f"총 {len(invoices)}개의 송장번호를 읽었습니다.")
            return invoices
            
        except Exception as e:
            logger.error(f"엑셀 파일 읽기 실패: {e}")
            raise
    
    def validate_data(self) -> bool:
        """
        데이터 검증
        
        Returns:
            검증 성공 여부
        """
        try:
            invoices = self.read_invoices()
            if len(invoices) == 0:
                logger.warning("읽은 송장번호가 없습니다.")
                return False
            return True
        except Exception as e:
            logger.error(f"데이터 검증 실패: {e}")
            return False

